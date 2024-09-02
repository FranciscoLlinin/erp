# -*- coding:utf-8 -*-
from odoo import models, fields, api
from datetime import datetime as dt
from datetime import date
import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

_logger = logging.getLogger(__name__)

class ProyectoMvg(models.Model):
    _name = "project.mvg"
    _rec_name = 'project_id'

    currency_id = fields.Many2one('res.currency', string='Currency')
    project_id = fields.Many2one('project.project', string='Proyecto')

    costo_directo = fields.Monetary(string="Costo directo")
    fecha_inicio = fields.Datetime(string="Fecha de inicio")
    fecha_termino = fields.Datetime(string="Fecha de término")
    duracion = fields.Float(string="Duración en días laborables", digits=(16, 2))
    precio_venta = fields.Float(string="Precio de venta", digits=(16, 2), compute='_precio_venta')

    budget_at_completition = fields.Monetary(string="Presupuestado")
    planned_value = fields.Monetary(string="Valor planeado")
    earned_value = fields.Monetary(string="Valor ganado")
    actual_cost = fields.Monetary(string="Costo actual")
    ev_record_date = fields.Datetime(string="Fecha de corte")

    # CV = EV - AC
    cost_variance = fields.Monetary(string="Variación del costo", compute='_cost_variance')
    # SV = EV - PV
    scheduled_variance = fields.Monetary(string="Variación del programa", compute='_scheduled_variance')
    # CPI = EV / AC
    cost_performance_index = fields.Monetary(string="Índice desempeño costo", compute='_cost_performance_index')
    # SPI = EV / PV
    scheduled_performance_index = fields.Monetary(string="Índice desempeño programa",
                                                  compute='_scheduled_performance_index')
    # CSI = CPI * SPI. Este indicador nos da una relación entre el costo y el cronograma y así saber que posibilidades
    #                  tenemos de recuperar nuestro proyecto.
    # CSI > 0,9	Proyecto OK
    # CSI Entre 0,8 y 0,9	Hay posibilidades de arreglarlo
    # CSI < 0.8	Lo más probable es que no se arregle.
    cost_scheduled_index = fields.Monetary(string="Índice costo-programación", compute='_cost_scheduled_index')
    # EAC = BAC / CPI
    estimated_at_completion = fields.Monetary(string="Estimado a la terminación", compute='_estimated_at_completion')
    # ETC = EAC - AC
    estimated_to_completion = fields.Monetary(string="Estimado para completar", compute='_estimated_to_completion')
    # VAC = BAC - EAC
    variance_at_completion = fields.Monetary(string="Variación a la terminación", compute='_variance_at_completion')
    # ED = Days / SPI
    estimated_days = fields.Float(string="Días estimados", compute='_estimated_days')
    # EED = FI + ED
    estimated_end_date = fields.Datetime(string="Fecha de término estimada", compute='_estimated_end_date')

    # AP = EV * 100 / BAC
    advance_percent = fields.Float(string="% Avance", compute='_advance_percent')

    # EP = AC * 100 / BAC
    estimated_percent = fields.Float(string="% Estimado", compute='_estimated_percent')

    invoiced = fields.Float(string="Facturado")
    paid = fields.Float(string="Pagado")
    paid_difference = fields.Float(string="Diferencia pagado", compute='_paid_difference')
    obs = fields.Char(string="Observaciones")
    # PP = paid * 100 / invoiced
    paid_percent = fields.Float(string="% Pagado", compute='_paid_percent')

    project_state = fields.Selection([
        ('Ok', 'Ok'),
        ('Detenido', 'Detenido'),
        ('Critico', 'Crítico'),
        ('Detalles', 'Detalles')
    ], string='Estado', default='Ok', required=True, editable=True)

    # @api.one
    def _precio_venta(self):
        for record in self:
            record.precio_venta = record.costo_directo #self.project_id.precio_venta

    # @api.one
    def _cost_variance(self):
        for record in self:
            record.cost_variance = record.earned_value - record.actual_cost

    # @api.one
    def _scheduled_variance(self):
        for record in self:
            record.scheduled_variance = record.earned_value - record.planned_value

    # @api.one
    def _cost_performance_index(self):
        for record in self:
            if record.actual_cost != 0:
                record.cost_performance_index = record.earned_value / record.actual_cost
            else:
                record.cost_performance_index = 0

    # @api.one
    def _scheduled_performance_index(self):
        for record in self:
            if record.planned_value != 0:
                record.scheduled_performance_index = record.earned_value / record.planned_value
            else:
                record.scheduled_performance_index = 0

    # @api.one
    def _cost_scheduled_index(self):
        for record in self:
            record.cost_scheduled_index = record.cost_performance_index * record.scheduled_performance_index

    # @api.one
    def _estimated_at_completion(self):
        for record in self:
            if record.cost_performance_index != 0:
                record.estimated_at_completion = record.budget_at_completition / record.cost_performance_index
            else:
                record.estimated_at_completion = 0

    # @api.one
    def _estimated_to_completion(self):
        for record in self:
            record.estimated_to_completion = record.estimated_at_completion - record.actual_cost

    # @api.one
    def _variance_at_completion(self):
        for record in self:
            record.variance_at_completion = record.budget_at_completition - record.estimated_at_completion

    # @api.one
    def _estimated_days(self):
        for record in self:
            if record.scheduled_performance_index != 0:
                # inicio = dt.strptime(self.fecha_inicio, '%Y-%m-%d %H:%M:%S')
                # termino = dt.strptime(self.fecha_termino, '%Y-%m-%d %H:%M:%S')
                # duration = (termino - inicio).days
                duration = (record.fecha_termino - record.fecha_inicio).days
                record.estimated_days = duration / record.scheduled_performance_index
            else:
                record.estimated_days = 0

    # @api.one
    def _estimated_end_date(self):
        for record in self:
            # inicio = dt.strptime(self.fecha_inicio, '%Y-%m-%d %H:%M:%S')
            # self.estimated_end_date = inicio + datetime.timedelta(days=self.estimated_days)
            record.estimated_end_date = record.fecha_inicio + datetime.timedelta(days=record.estimated_days)

    # @api.one
    def _advance_percent(self):
        for record in self:
            if record.budget_at_completition != 0:
                record.advance_percent = record.earned_value * 100 / record.budget_at_completition
            else:
                record.advance_percent = 0

    # @api.one
    def _estimated_percent(self):
        for record in self:
            if record.budget_at_completition != 0:
                record.estimated_percent = record.actual_cost * 100 / record.budget_at_completition
            else:
                record.estimated_percent = 0

    # @api.one
    def _paid_percent(self):
        for record in self:
            if record.invoiced != 0:
                record.paid_percent = record.paid * 100 / record.invoiced
            else:
                record.paid_percent = 0

    # @api.one
    def _paid_difference(self):

        for record in self:
            record.paid_difference = record.invoiced - record.paid

    def some_condition(self):
        return True

    def actualiza_facturas_pagos(self):

        active_ids = self._context.get('active_ids', [])
        mvgs = self.browse(active_ids)

        for mvg in mvgs:
            facturas_lineas = self.env['account.move.line'].search(['|', ('analytic_account_id.code', '=', 'OBRA'),
                                                                       ('analytic_account_id.code', '=', 'SERVICIOS'),
                                                                       ('analytic_account_id.name', '=', mvg.project_id.name),
                                                                       ('move_id.move_type', '=', 'out_invoice'),
                                                                       ('exclude_from_invoice_tab', '=', False)
                                                                       ])
            dicc_facturas_con_iva = {}
            dicc_facturas_sin_iva = {}
            dicc_pagos = {}
            facturas_procesadas = []
            for linea in facturas_lineas:
                obra = linea.analytic_account_id.name

                factura = linea.move_id
                _logger.debug(f"Factura: {factura.name}")

                if factura.id not in facturas_procesadas:
                    facturas_procesadas.append(factura.id)
                    if obra in dicc_facturas_sin_iva:
                        dicc_facturas_sin_iva[obra] += factura.amount_untaxed_signed
                        dicc_facturas_con_iva[obra] += factura.amount_total
                    else:
                        dicc_facturas_sin_iva[obra] = factura.amount_untaxed_signed
                        dicc_facturas_con_iva[obra] = factura.amount_total

                pagos = factura._get_reconciled_invoices_partials()

                if pagos and len(pagos) > 0:
                    pago = pagos[0]
                    apr = pago[0]
                    aml = pago[2]
                    amount = sum(float(s[1]) for s in pagos)
                    amls = apr.debit_move_id.move_id.line_ids

                    #Busco el que tiene tax_base_amount > 0
                    tba = next(a for a in amls if a.tax_base_amount > 0)

                    #Busco en apr el recién encontrado y ese es su importe de iva que tengo que descontar
                    found = self.env['account.partial.reconcile'].search([('credit_move_id', '=', tba.id)])
                    if found:
                        for f in found:
                            amount = round(amount - f.amount, 2)

                    if obra in dicc_pagos:
                        dicc_pagos[obra] += amount
                    else:
                        dicc_pagos[obra] = amount

            for obra in dicc_facturas_sin_iva:
                _logger.debug(f"Obra pagos: {obra}")
                # mvg = next((m for m in mvgs if m.project_id.name == obra), "none")
                pagado = dicc_pagos[obra] if obra in dicc_pagos else 0
                if mvg is 'none':
                    proyecto = self.env['project.project'].search([('analytic_account_id.name', '=', obra)])
                    if proyecto.id is not False:
                        inicio = proyecto.fecha_inicio if proyecto.fecha_inicio is True else date.today()
                        termino = proyecto.fecha_termino if proyecto.fecha_termino is True else date.today()
                        self.env['project.mvg'].create({
                            'project_id': proyecto.id,
                            'invoiced': dicc_facturas_sin_iva[obra],
                            'paid': pagado,
                            'fecha_inicio': inicio,
                            'fecha_termino': termino,
                            'obs': obra
                        })
                else:
                    mvg.write({
                        'invoiced': dicc_facturas_sin_iva[obra],
                        'paid': pagado
                    })

        _logger.debug(f"FIN PROCESO")

    def crea_dummy(self):
        inicio = datetime.datetime(2021, 1, 1)
        termino = datetime.datetime(2021, 12, 31)
        obra = 'OBRA X'
        proyecto = self.env['project.project'].search([('analytic_account_id.name', '=', obra)])
        self.env['project.mvg'].create({
            'project_id': proyecto.id,
            'fecha_inicio': inicio,
            'fecha_termino': termino,
            'obs': obra
        })

    def enviar_excel(self):
        # self.crea_dummy()
        # return
        self.genera_excel()
        data = {
            'type': 'ir.actions.act_url',
            'url': '/metodo_valor_ganado',
            'target': 'new',
        }
        return data

    def genera_excel(self):
        mvgs = self.get_mvgs()

        wb = Workbook()
        ws = wb.active
        ws.title = "Método de valor ganado"

        self.arma_titulos(ws)
        self.arma_excel(ws, mvgs)

        file_name = "valor_ganado.xlsx"
        directorio = '/var/lib/odoo/.local/share/Odoo/Ecosoft'
        wb.save(directorio + '/' + file_name)

        return {
            'name': 'Message',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'custom.pop.message',
            'target': 'new',
            'context': {'default_name': "Excel generado"}
        }

    def get_mvgs(self):
        # active_model = self.env.context.get('active_model', False)
        # active_ids = self.env.context.get('active_ids', [])
        # data_source = self.env['project.mvg'].browse(self._context.get('active_ids'))
        # xx = self.env.context  # ['active_ids']
        mvgs = self.env['project.mvg'].search([])
        return mvgs

    def arma_titulos(self,ws):
        ws['A1'] = "PROYECTO"
        ws['A1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['B1'] = "FECHA DE CORTE"
        ws['B1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['C1'] = "INICIO"
        ws['C1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['D1'] = "TÉRMINO"
        ws['D1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['E1'] = "PRESUPUESTADO"
        ws['E1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['F1'] = "VALOR PLANEADO"
        ws['F1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['G1'] = "VALOR GANADO"
        ws['G1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['H1'] = "COSTO ACUTAL"
        ws['H1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['I1'] = "VARIACIÓN DEL COSTO"
        ws['I1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['J1'] = "VARIACIÓN DEL PROGRAMA"
        ws['J1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['K1'] = "ÍNDICE DESEMPEÑO COSTO"
        ws['K1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['L1'] = "ÍNDICE DESEMPEÑO PROGRAMA"
        ws['L1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['M1'] = "INDICE COSTO-PROGRAMACIÓN"
        ws['M1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['N1'] = "ESTIMADO A LA TERMINACIÓN"
        ws['N1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['O1'] = "ESTIMADO PARA COMPLETAR"
        ws['O1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['P1'] = "VARIACIÓN A LA TERMINACIÓN"
        ws['P1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['Q1'] = "DÍAS ESTIMADOS"
        ws['Q1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['R1'] = "FECHA DE TERMINO ESTIMADA"
        ws['R1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['S1'] = "PRECIO DE VENTA"
        ws['S1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['T1'] = "FACTURADO"
        ws['T1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['U1'] = "PAGADO"
        ws['U1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['V1'] = "DIFERENCIA PAGADO"
        ws['V1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['W1'] = "% AVANCE"
        ws['W1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['X1'] = "% ESTIMADO"
        ws['X1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['Y1'] = "% PAGADO"
        ws['Y1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        ws['Z1'] = "ESTADO"
        ws['Z1'].font = Font(name='Calibri', size=10, color="094293", bold=True)
        return True

    def arma_excel(self,ws, mvgs):

        i=2
        for mvg in mvgs:
            proyecto = mvg.project_id.name
            ev_record_date = mvg.ev_record_date
            fecha_inicio = mvg.fecha_inicio
            fecha_termino = mvg.fecha_termino
            budget_at_completition = mvg.budget_at_completition
            planned_value = mvg.planned_value
            earned_value = mvg.earned_value
            actual_cost = mvg.actual_cost
            cost_variance = mvg.cost_variance
            scheduled_variance = mvg.scheduled_variance
            cost_performance_index = mvg.cost_performance_index
            scheduled_performance_index = mvg.scheduled_performance_index
            cost_scheduled_index = mvg.cost_scheduled_index
            estimated_at_completion = mvg.estimated_at_completion
            estimated_to_completion = mvg.estimated_to_completion
            variance_at_completion = mvg.variance_at_completion
            estimated_days = mvg.estimated_days
            estimated_end_date = mvg.estimated_end_date
            precio_venta = mvg.precio_venta
            invoiced = mvg.invoiced
            paid = mvg.paid
            paid_difference = mvg.paid_difference
            advance_percent = mvg.advance_percent
            estimated_percent = mvg.estimated_percent
            paid_percent = mvg.paid_percent
            project_state = mvg.project_state


            ws.cell(row=i, column=1, value=proyecto)
            ws.cell(row=i, column=2, value=ev_record_date)
            ws.cell(row=i, column=3, value=fecha_inicio)
            ws.cell(row=i, column=4, value=fecha_termino)
            ws.cell(row=i, column=5, value=budget_at_completition)
            ws.cell(row=i, column=6, value=planned_value)
            ws.cell(row=i, column=7, value=earned_value)
            ws.cell(row=i, column=8, value=actual_cost)
            ws.cell(row=i, column=9, value=cost_variance)
            ws.cell(row=i, column=10, value=scheduled_variance)
            ws.cell(row=i, column=11, value=cost_performance_index)
            ws.cell(row=i, column=12, value=scheduled_performance_index)
            ws.cell(row=i, column=13, value=cost_scheduled_index)
            ws.cell(row=i, column=14, value=estimated_at_completion)
            ws.cell(row=i, column=15, value=estimated_to_completion)
            ws.cell(row=i, column=16, value=variance_at_completion)
            ws.cell(row=i, column=17, value=estimated_days)
            ws.cell(row=i, column=18, value=estimated_end_date)
            ws.cell(row=i, column=19, value=precio_venta)
            ws.cell(row=i, column=20, value=invoiced)
            ws.cell(row=i, column=21, value=paid)
            ws.cell(row=i, column=22, value=paid_difference)
            ws.cell(row=i, column=23, value=advance_percent)
            ws.cell(row=i, column=24, value=estimated_percent)
            ws.cell(row=i, column=25, value=paid_percent)
            ws.cell(row=i, column=26, value=project_state)
            i=i+1


        for col in ws.iter_cols(min_col=1, max_col=26, min_row=2, max_row=i):
            for cell in col:
                # cell.fill = PatternFill(start_color="E5E7E9", end_color="E5E7E9", fill_type='solid')
                # cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=8, color=colors.BLACK, bold=False)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 15
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 15
        ws.column_dimensions["O"].width = 15
        ws.column_dimensions["P"].width = 15
        ws.column_dimensions["Q"].width = 15
        ws.column_dimensions["R"].width = 15
        ws.column_dimensions["S"].width = 15
        ws.column_dimensions["T"].width = 15
        ws.column_dimensions["U"].width = 15
        ws.column_dimensions["V"].width = 15
        ws.column_dimensions["W"].width = 15
        ws.column_dimensions["X"].width = 15
        ws.column_dimensions["Y"].width = 15
        ws.column_dimensions["Z"].width = 15

        return True

    def actualiza_facturas_pagos_todos(self):
        for record in self:
            print('Este es un registro')
            print(record)
        invoice_ids = self._context.get('active_ids', [])
        context = self._context
        datas = {'ids': context.get('active_ids', [])}
        context = dict(self._context or {})
        active_ids = context.get('active_ids', [])
        ids = self.ids
        active_model = self.env.context.get('active_model', False)
        active_ids = self.env.context.get('active_ids', [])
        src_list_ids = self.env.context.get('active_ids')
        active_ids = self._context.get('active_ids')
        data_source = self.env['project.mvg'].browse(self._context.get('active_ids'))
        xx = self.env.context  # ['active_ids']

        mvgs = self.env['project.mvg'].search([])


        facturas_lineas = self.env['account.move.line'].search(['|', ('analytic_account_id.code', '=', 'OBRA'),
                                                                   ('analytic_account_id.code', '=', 'SERVICIOS'),
                                                                   ('move_id.move_type', '=', 'out_invoice'),
                                                                   ('exclude_from_invoice_tab', '=', False)
                                                                   ])
        dicc_facturas_con_iva = {}
        dicc_facturas_sin_iva = {}
        dicc_pagos = {}
        facturas_procesadas = []
        for linea in facturas_lineas:
            obra = linea.analytic_account_id.name

            mvg = next((m for m in mvgs if m.project_id.name == obra), "none")

            factura = linea.move_id
            _logger.debug(f"Factura: {factura.name}")

            if factura.id not in facturas_procesadas:
                facturas_procesadas.append(factura.id)
                if obra in dicc_facturas_sin_iva:
                    dicc_facturas_sin_iva[obra] += factura.amount_untaxed_signed
                    dicc_facturas_con_iva[obra] += factura.amount_total
                else:
                    dicc_facturas_sin_iva[obra] = factura.amount_untaxed_signed
                    dicc_facturas_con_iva[obra] = factura.amount_total

            pagos = factura._get_reconciled_invoices_partials()

            # for pago in pagos:
            if pagos and len(pagos) > 0:
                pago = pagos[0]
                apr = pago[0]
                aml = pago[2]
                amount = sum(float(s[1]) for s in pagos)
                # aml = self.env['account.move.line'].search([('id', '=', apr.debit_move_id)])
                # aml = apr.debit_move_id
                # amls = self.env['account.move.line'].search([('ref', '=', apr.debit_move_id.move_name)])
                amls = apr.debit_move_id.move_id.line_ids

                #Busco el que tiene tax_base_amount > 0

                tba = next(a for a in amls if a.tax_base_amount > 0)

                #Busco en apr el recién encontrado y ese es su importe de iva que tengo que descontar
                found = self.env['account.partial.reconcile'].search([('credit_move_id', '=', tba.id)])
                if found:
                    for f in found:
                        amount = round(amount - f.amount, 2)

                if obra in dicc_pagos:
                    dicc_pagos[obra] += amount
                else:
                    dicc_pagos[obra] = amount

        for obra in dicc_facturas_sin_iva:
            _logger.debug(f"Obra pagos: {obra}")
            mvg = next((m for m in mvgs if m.project_id.name == obra), "none")
            pagado = dicc_pagos[obra] if obra in dicc_pagos else 0
            if mvg is 'none':
                proyecto = self.env['project.project'].search([('analytic_account_id.name', '=', obra)])
                if proyecto.id is not False:
                    inicio = proyecto.fecha_inicio if proyecto.fecha_inicio is True else date.today()
                    termino = proyecto.fecha_termino if proyecto.fecha_termino is True else date.today()
                    self.env['project.mvg'].create({
                        'project_id': proyecto.id,
                        'invoiced': dicc_facturas_sin_iva[obra],
                        'paid': pagado,
                        'fecha_inicio': inicio,
                        'fecha_termino': termino,
                        'obs': obra
                    })
            else:
                mvg.write({
                    'invoiced': dicc_facturas_sin_iva[obra],
                    'paid': pagado
                })

        _logger.debug(f"FIN PROCESO")
